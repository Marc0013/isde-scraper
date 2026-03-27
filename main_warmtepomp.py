"""
ISDE Warmtepomp Scraper - volledig werkende pipeline
Haalt alle warmtepomp-meldcodes op van RVO.nl en maakt een vergelijkingstabel
met subsidies voor 2024, 2025 en huidig tarief.

Gebruik:
    python main_warmtepomp.py               # volledige run (~2.973 pagina's)
    python main_warmtepomp.py --limit 50    # test met 50 pagina's
    python main_warmtepomp.py --skip-fetch  # sla downloaden over, hergebruik parsed data
    python main_warmtepomp.py --only-merge  # maak alleen CSV/JSON opnieuw uit bestaande data
"""

import sys
import time
import argparse
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).parent))

from src.utils.logger import setup_logger
from src.utils.file_utils import save_json, load_json, now_iso
from src.scraper.sitemap_parser import fetch_warmtepomp_urls
from src.scraper.domain_filter import filter_url
from src.parser.warmtepomp_detail_parser import parse_warmtepomp_detail

logger = setup_logger("warmtepomp", log_dir="logs")

HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; ISDE-subsidie-scraper/1.0; research)",
    "Accept-Language": "nl-NL,nl;q=0.9",
}
DELAY_SECONDS = 1.5
REQUEST_TIMEOUT = 20
TARGET_YEARS = {2024, 2025, 2026}

PARSED_DIR   = Path("data/parsed/warmtepompen")
NORM_DIR     = Path("data/normalized")
COMPARE_DIR  = Path("data/comparison")
VISITED_FILE = Path("data/raw/metadata/warmtepomp_visited.json")


# ─────────────────────────────────────────────────────────────────────────────
# Stap 1: Discovery
# ─────────────────────────────────────────────────────────────────────────────

def step_discover(limit=None) -> list[dict]:
    logger.info("STAP 1: DISCOVERY via sitemap")
    urls = fetch_warmtepomp_urls(limit=limit)
    logger.info(f"  {len(urls)} warmtepomp URLs gevonden")
    return urls


# ─────────────────────────────────────────────────────────────────────────────
# Stap 2: Fetch + parse
# ─────────────────────────────────────────────────────────────────────────────

def step_fetch_and_parse(discovered: list[dict]) -> list[dict]:
    """
    Download elke detailpagina en parseer hem direct.
    Slaat individuele parsed records op als JSON.
    Herstartbaar: reeds geparseerde URLs worden overgeslagen.
    """
    logger.info("STAP 2: FETCH + PARSE")
    PARSED_DIR.mkdir(parents=True, exist_ok=True)

    visited = set(load_json(VISITED_FILE) or [])
    session = requests.Session()
    session.headers.update(HEADERS)

    parsed_records = []
    skipped = 0
    failed = 0

    for i, item in enumerate(discovered, 1):
        url = item["url"]

        # Domeincheck (hard)
        if not filter_url(url):
            logger.warning(f"Geblokkeerd (extern domein): {url}")
            continue

        # Al eerder verwerkt?
        if url in visited:
            # Laad bestaand parsed bestand
            slug = item.get("slug", "")
            cached = _load_parsed_cache(slug)
            if cached:
                parsed_records.append(cached)
                skipped += 1
                continue

        logger.info(f"[{i}/{len(discovered)}] {url}")
        time.sleep(DELAY_SECONDS)

        try:
            r = session.get(url, timeout=REQUEST_TIMEOUT)
            if r.status_code == 404:
                logger.warning(f"  404 - overgeslagen")
                visited.add(url)
                continue
            if r.status_code != 200:
                logger.warning(f"  HTTP {r.status_code} - overgeslagen")
                failed += 1
                continue

            parsed = parse_warmtepomp_detail(
                html=r.text,
                source_url=url,
                source_file="",
            )

            if parsed and parsed.get("meldcode"):
                _save_parsed_record(parsed, item.get("slug", ""))
                parsed_records.append(parsed)
                visited.add(url)

                jaren = [sb["jaar"] for sb in parsed.get("subsidiebedragen", []) if sb.get("jaar")]
                logger.info(
                    f"  {parsed['meldcode']} | {parsed.get('fabrikant','')} "
                    f"| jaren: {jaren} | confidence: {parsed.get('confidence')}"
                )
            else:
                logger.warning(f"  Geen bruikbare data gevonden")
                failed += 1

        except requests.exceptions.Timeout:
            logger.warning(f"  Timeout - overgeslagen")
            failed += 1
        except Exception as e:
            logger.error(f"  Fout: {e}")
            failed += 1

        # Periodiek visited opslaan
        if i % 50 == 0:
            save_json(list(visited), VISITED_FILE)
            logger.info(f"  Voortgang: {i}/{len(discovered)} verwerkt, {len(parsed_records)} OK, {failed} mislukt")

    save_json(list(visited), VISITED_FILE)
    logger.info(f"Fetch/parse klaar: {len(parsed_records)} OK, {skipped} hergebruikt, {failed} mislukt")
    return parsed_records


# ─────────────────────────────────────────────────────────────────────────────
# Stap 3: Normaliseer per jaar
# ─────────────────────────────────────────────────────────────────────────────

def _is_tweede(label: str) -> bool:
    """True als het label een '2e warmtepomp' tarief aangeeft."""
    return "2e" in label.lower() or "tweede" in label.lower()


def step_normalize(parsed_records: list[dict]) -> dict:
    """
    Splits parsed records op jaar.
    Onderscheidt 'regulier' en '2e warmtepomp' subsidies per jaar.
    """
    logger.info("STAP 3: NORMALISEREN per jaar")

    # Per jaar twee buckets: regulier en 2e warmtepomp
    by_year = {jaar: {"regulier": [], "tweede": []} for jaar in TARGET_YEARS}

    for rec in parsed_records:
        base = {
            "meldcode":         rec.get("meldcode"),
            "fabrikant":        rec.get("fabrikant"),
            "model":            rec.get("model"),
            "vermogen_kw":      rec.get("vermogen_kw"),
            "naam_koudemiddel": rec.get("naam_koudemiddel"),
            "gwp":              rec.get("gwp"),
            "categorie":        rec.get("categorie"),
            "bron_url":         rec.get("source_url"),
            "confidence":       rec.get("confidence", 0.85),
        }
        for sb in rec.get("subsidiebedragen", []):
            jaar = sb.get("jaar")
            if jaar not in TARGET_YEARS:
                continue
            bucket = "tweede" if _is_tweede(sb.get("label_origineel", "")) else "regulier"
            by_year[jaar][bucket].append({
                **base,
                "subsidiebedrag": sb.get("bedrag"),
                "label_origineel": sb.get("label_origineel", ""),
            })

    for jaar, buckets in by_year.items():
        out = NORM_DIR / str(jaar) / f"warmtepompen_{jaar}.json"
        out.parent.mkdir(parents=True, exist_ok=True)
        save_json(buckets["regulier"], out)
        logger.info(f"  {jaar}: {len(buckets['regulier'])} regulier, {len(buckets['tweede'])} 2e warmtepomp")

    return by_year


# ─────────────────────────────────────────────────────────────────────────────
# Stap 4: Merge 2024 + 2025 → vergelijkingstabel
# ─────────────────────────────────────────────────────────────────────────────

def step_merge(by_year: dict) -> list[dict]:
    """
    Koppel records van 2024, 2025 en 2026 op meldcode.
    Eén rij per product, met aparte kolommen per jaar.
    2026 heeft ook een '2e warmtepomp' kolom.
    """
    logger.info("STAP 4: MERGEN 2024 + 2025 + 2026")

    import re
    def nk(v):
        return re.sub(r"[\s\-_]", "", str(v).upper()) if v else None

    def make_index(records):
        return {nk(r["meldcode"]): r for r in records if r.get("meldcode")}

    idx = {
        jaar: {
            "regulier": make_index(by_year[jaar]["regulier"]),
            "tweede":   make_index(by_year[jaar]["tweede"]),
        }
        for jaar in TARGET_YEARS
    }

    all_keys = set()
    for jaar in TARGET_YEARS:
        all_keys |= set(idx[jaar]["regulier"]) | set(idx[jaar]["tweede"])

    today = now_iso()[:10]
    comparison = []

    for key in sorted(all_keys):
        r24  = idx[2024]["regulier"].get(key)
        r25  = idx[2025]["regulier"].get(key)
        r26  = idx[2026]["regulier"].get(key)
        r26b = idx[2026]["tweede"].get(key)
        base = r26 or r25 or r24 or r26b

        aanwezig = [y for y, r in [(2024, r24), (2025, r25), (2026, r26)] if r]
        if len(aanwezig) >= 2:
            match_methode = "meldcode"
        elif aanwezig:
            match_methode = f"alleen_{aanwezig[0]}"
        else:
            match_methode = "alleen_2026_2e"

        row = {
            "meldcode":                base["meldcode"],
            "fabrikant":               base.get("fabrikant"),
            "model":                   base.get("model"),
            "vermogen_kw":             base.get("vermogen_kw"),
            "naam_koudemiddel":        base.get("naam_koudemiddel"),
            "gwp":                     base.get("gwp"),
            "categorie":               base.get("categorie"),
            # Subsidies per jaar
            "subsidiebedrag_2024":     r24["subsidiebedrag"]  if r24  else None,
            "subsidiebedrag_2025":     r25["subsidiebedrag"]  if r25  else None,
            "subsidiebedrag_2026":     r26["subsidiebedrag"]  if r26  else None,
            "subsidiebedrag_2e_2026":  r26b["subsidiebedrag"] if r26b else None,
            # Bronnen
            "bron_url":                base.get("bron_url"),
            # Kwaliteit
            "confidence_2024":         r24["confidence"]  if r24  else None,
            "confidence_2025":         r25["confidence"]  if r25  else None,
            "confidence_2026":         r26["confidence"]  if r26  else None,
            "match_methode":           match_methode,
            "datum_opgehaald":         today,
        }
        comparison.append(row)

    logger.info(f"  {len(comparison)} vergelijkingsrecords")
    for jaar in [2024, 2025, 2026]:
        col = f"subsidiebedrag_{jaar}"
        n = sum(1 for r in comparison if r.get(col) is not None)
        logger.info(f"  {jaar}: {n} records met subsidiebedrag")
    n2e = sum(1 for r in comparison if r.get("subsidiebedrag_2e_2026") is not None)
    logger.info(f"  2026 (2e warmtepomp): {n2e} records")

    return comparison


# ─────────────────────────────────────────────────────────────────────────────
# Stap 5: Opslaan als JSON + CSV
# ─────────────────────────────────────────────────────────────────────────────

def step_save(comparison: list[dict]) -> None:
    logger.info("STAP 5: OPSLAAN")
    COMPARE_DIR.mkdir(parents=True, exist_ok=True)

    if not comparison:
        logger.warning("Geen records om op te slaan")
        return

    json_path = COMPARE_DIR / "isde_warmtepompen_vergelijking.json"
    csv_path  = COMPARE_DIR / "isde_warmtepompen_vergelijking.csv"
    xlsx_path = COMPARE_DIR / "isde_warmtepompen_vergelijking.xlsx"

    save_json(comparison, json_path)
    logger.info(f"  JSON: {json_path} ({len(comparison)} records)")

    import pandas as pd

    # Kolomvolgorde: productinfo | subsidies per jaar | meta
    col_order = [
        "meldcode", "fabrikant", "model", "vermogen_kw",
        "naam_koudemiddel", "gwp", "categorie",
        "subsidiebedrag_2024",
        "subsidiebedrag_2025",
        "subsidiebedrag_2026",
        "subsidiebedrag_2e_2026",
        "bron_url",
        "confidence_2024", "confidence_2025", "confidence_2026",
        "match_methode", "datum_opgehaald",
    ]
    df = pd.DataFrame(comparison)
    # Zorg dat alle verwachte kolommen aanwezig zijn
    for col in col_order:
        if col not in df.columns:
            df[col] = None
    df = df[col_order]

    # ── CSV met puntkomma (werkt direct in NL Excel) ──
    df.to_csv(str(csv_path), index=False, encoding="utf-8-sig", sep=";")
    logger.info(f"  CSV:  {csv_path} ({len(df)} rijen, puntkomma-gescheiden)")

    # ── Excel met opmaak ──
    try:
        with pd.ExcelWriter(str(xlsx_path), engine="openpyxl") as writer:
            df.to_excel(writer, index=False, sheet_name="Warmtepompen")
            ws = writer.sheets["Warmtepompen"]

            # Kolombreedtes aanpassen
            col_widths = {
                "meldcode": 12, "fabrikant": 22, "model": 30,
                "vermogen_kw": 12, "naam_koudemiddel": 16, "gwp": 8,
                "categorie": 20,
                "subsidiebedrag_2024": 22, "subsidiebedrag_2025": 22,
                "subsidiebedrag_2026": 22, "subsidiebedrag_2e_2026": 24,
                "bron_url": 22,
                "confidence_2024": 14, "confidence_2025": 14, "confidence_2026": 14,
                "match_methode": 14, "datum_opgehaald": 16,
            }
            for i, col in enumerate(col_order, 1):
                ws.column_dimensions[
                    ws.cell(1, i).column_letter
                ].width = col_widths.get(col, 14)

            # Koptekstrij vet + lichtgrijs
            from openpyxl.styles import Font, PatternFill, Alignment
            header_fill = PatternFill("solid", fgColor="D9D9D9")
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal="center")

            # Kleuropmaak per kolomtype
            yellow_hdr = PatternFill("solid", fgColor="FFD966")   # geel header subsidie
            yellow_cel = PatternFill("solid", fgColor="FFF2CC")   # geel cel subsidie
            orange_hdr = PatternFill("solid", fgColor="F4B942")   # oranje header 2e wp
            orange_cel = PatternFill("solid", fgColor="FDE9C4")   # oranje cel 2e wp
            light_blue = PatternFill("solid", fgColor="EEF4FB")   # blauw afwisseling

            subsidie_cols = [i + 1 for i, c in enumerate(col_order) if "subsidiebedrag" in c and "2e" not in c]
            tweede_cols   = [i + 1 for i, c in enumerate(col_order) if "2e_2026" in c]

            for col_idx in subsidie_cols:
                ws.cell(1, col_idx).fill = yellow_hdr
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row_idx, col_idx).fill = yellow_cel

            for col_idx in tweede_cols:
                ws.cell(1, col_idx).fill = orange_hdr
                for row_idx in range(2, len(df) + 2):
                    ws.cell(row_idx, col_idx).fill = orange_cel

            alle_subsidie = set(subsidie_cols) | set(tweede_cols)
            for row_idx in range(2, len(df) + 2):
                if row_idx % 2 == 0:
                    for col_idx in range(1, len(col_order) + 1):
                        if col_idx not in alle_subsidie:
                            ws.cell(row_idx, col_idx).fill = light_blue

            # Bevroren bovenste rij
            ws.freeze_panes = "A2"

        logger.info(f"  XLSX: {xlsx_path} ({len(df)} rijen, {len(col_order)} kolommen)")

    except ImportError:
        logger.warning("openpyxl niet beschikbaar, alleen CSV en JSON opgeslagen")


# ─────────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────────

def _save_parsed_record(parsed: dict, slug: str) -> None:
    safe_slug = slug.replace("/", "_")[:80]
    path = PARSED_DIR / f"{safe_slug}.json"
    save_json(parsed, path)


def _load_parsed_cache(slug: str) -> dict | None:
    safe_slug = slug.replace("/", "_")[:80]
    path = PARSED_DIR / f"{safe_slug}.json"
    return load_json(path)


def _load_all_parsed() -> list[dict]:
    """Laad alle eerder geparseerde records vanuit parsed map."""
    records = []
    for f in sorted(PARSED_DIR.glob("*.json")):
        if f.name in ("test_warmtepomp_records.json", "test_summary.json"):
            continue
        data = load_json(f)
        if data and isinstance(data, dict) and data.get("meldcode"):
            records.append(data)
    logger.info(f"Geladen uit cache: {len(records)} records")
    return records


# ─────────────────────────────────────────────────────────────────────────────
# Entrypoint
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="ISDE Warmtepomp Scraper")
    parser.add_argument("--limit", type=int, default=None,
                        help="Max aantal URLs (standaard: alle ~2973)")
    parser.add_argument("--skip-fetch", action="store_true",
                        help="Sla downloaden over, gebruik bestaande parsed data")
    parser.add_argument("--only-merge", action="store_true",
                        help="Sla alles over, maak alleen CSV/JSON opnieuw")
    args = parser.parse_args()

    logger.info("=" * 60)
    logger.info("ISDE WARMTEPOMP SCRAPER")
    logger.info("=" * 60)

    if args.only_merge:
        parsed_records = _load_all_parsed()
    elif args.skip_fetch:
        parsed_records = _load_all_parsed()
    else:
        discovered = step_discover(limit=args.limit)
        parsed_records = step_fetch_and_parse(discovered)

    by_year = step_normalize(parsed_records)
    comparison = step_merge(by_year)
    step_save(comparison)

    logger.info("=" * 60)
    logger.info("KLAAR")
    logger.info(f"Resultaat: data/comparison/isde_warmtepompen_vergelijking.csv")
    logger.info("=" * 60)


if __name__ == "__main__":
    main()
